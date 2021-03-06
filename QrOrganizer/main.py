import cv2
import numpy as np
import pyzbar.pyzbar as pyzbar
from openpyxl import load_workbook

def qrcode():
    flag = 0
    cap = cv2.VideoCapture(0)
    font = cv2.FONT_HERSHEY_PLAIN

    while True:
        _, frame = cap.read()

        decodedObjects = pyzbar.decode(frame)

        cv2.imshow("Frame", frame)

        for obj in decodedObjects:
            h = int(obj.data) + 1
            wb = load_workbook(filename = 'QrOrganizer.xlsx')
            sheet=wb['Planilha1']
            wb.active
            print(sheet.cell(row=h,column=2).value)
            img = (sheet.cell(row=h, column=3).value)
            img = cv2.imread(img, 1)
            cv2.imshow('', img)
            cv2.waitKey(0)
            cv2.destroyAllWindows()
            flag = 1
            break

        key = cv2.waitKey(5)
        if key == 27 or flag == 1:
            break

def search_string():
    print('Digite o componente/equipamento que deseja buscar: ')
    eqp = input()

    wb = load_workbook(filename = 'QrOrganizer.xlsx')
    sheet = wb['Planilha1']
    wb.active

    for row in range(2,sheet.max_row+1):  
        for column in "B": 
            cell_name = "{}{}".format(column, row)
            if(sheet[cell_name].value == eqp):
                img = (sheet.cell(row=row, column=3).value)
                img = cv2.imread(img, 1)
                cv2.imshow('', img)
                cv2.waitKey(0)
                cv2.destroyAllWindows()
                break


while(True):
    print('Escolha a funcao utilizada:\n1)QR Code\n2)Strings\n3)Parar')
    option = int(input())

    if option == 1:
        qrcode()
        cv2.destroyAllWindows()

    elif option == 2:
        search_string()
        cv2.destroyAllWindows()

    elif option == 3:
        break
